# -*- coding: utf-8 -*-
"""
Created on Wed Dec 19 18:32:51 2018
@author: PC
"""
import csv
import wave
from openpyxl.workbook import Workbook
import wave
import pylab as pl
import numpy as np
from numpy import reshape
from matplotlib.font_manager import FontProperties
import os

'''
def readcsv(filename):
    fileobj=open(filename,"r")
    reader=csv.reader(fileobj)
    column=[row[0] for row in reader]
    firstcolumn=list(map(float,column))
    return firstcolumn
    
'''
#读取数据
def openwav(filepath):
    f = wave.open(filepath, "rb")
    
    #读取格式信息
    #(nchannels, sampwidth, framerate, nframes, comptype, compname)
    params = f.getparams()
    nchannels, sampwidth, framerate, nframes = params[:4]
    print('声道数:',nchannels)
    print('量化位数:',sampwidth)
    print('采样频率:',framerate)
    print('采样点数:',nframes)
    
    #读取波形数据
    str_data = f.readframes(nframes)
    f.close()
    
    #将波形数据转换为数组
    wave_data = np.fromstring(str_data, dtype=np.short)
    wave_data.shape = -1,2
    wave_data = wave_data.T
    time = np.arange(0, nframes) * (1.0 / framerate)
    max_num=max(wave_data[0])
    print("最大值为：",max_num)
    wave_da=wave_data[0]/max_num
    #print(max(wave_da))
    return wave_da

def getshuju(firstcolumn):
    m=0
    n=0
    order={}
    for i in range(0,len(firstcolumn)-1):
        if(i<n):
            continue
        else:
            if(abs(firstcolumn[i])>0.85):
                m+=1
                if(i+2760<len(firstcolumn)):
                    order[m]=firstcolumn[i-240:i+2760]
                else:
                    order[m]=firstcolumn[i-240:len(firstcolumn)]
                n=i+2760                       
    print(len(order.keys()))
    #print(order[117])
    return order

#写入excle
def writecsv(dict,nanm):
    wb=Workbook()
    sheet=wb.active
    m=int(len(dict.keys()))
    n=len(dict[1])
    for hang in range(1,m):
        for lie in range(1,n+1):
            sheet.cell(row=hang,column=lie).value=dict[hang][lie-1]
    wb.save(nanm+".xlsx")
#读取excle文件  
def readexcle(filename):
    wb= Workbook()
    sheet=wb.active
    rows=sheet.max_row
    columns=sheet.max_column
    strdata=[]
    m=[]
    for i in range(1,rows):
        for j in range(1,columns):
            m.append(sheet.cell(row=i,column=j).value)
        strdata.append(m)
    
    
#读波形文件并转为数组
def readwave(filename):
    waveflie=wave.open(filename,"rb")
    params=waveflie.getparams()
    nchannels, sampwidth, framerate, nframes = params[:4]
    #读取数字节
    str_data = f.readframes(nframes)
    f.close()
    #将波形数据转换为数组
    wave_data = np.fromstring(str_data, dtype=np.short)
    wave_data.shape = -1,2  #wave_data[0]表示左声道
    wave_data = wave_data.T
    time = np.arange(0, nframes) * (1.0 / framerate)

#返回文件夹名称数组
def getdir(dirs):
    dir_=os.listdir(dirs)
    return dir_
  
#filename="3.yuanshi.csv"
#list1=readcsv(filename)
#主程序入口
if __name__ == '__main__':
    z=r"C:\Users\PC\Desktop\123"
    directory=getdir(z)
    for i in directory:
        data=openwav(z+"\\"+i)
        array=getshuju(data)
        writecsv(array,i.split(".")[0])
        print(i+"完成切割")
    


